import os
import sys
import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font


def format_files(in_dir, out_dir):
    for i, f in enumerate(os.listdir(in_dir)):
        if os.path.isdir(f"{in_dir}/{f}"):
            format_files(f"{in_dir}/{f}", out_dir)
        else:
            if ".xlsx" in f:
                format_file(f"{in_dir}/{f}", out_dir)


thin = Side(border_style="thin")
def format_file(file_path, out_dir):
    wb = load_workbook(file_path)
    ws = wb.active
    customer_cell = ws['B7']

    ws.delete_cols(8, 2)
    ws.delete_cols(2, 1)

    last_table_row = find_last_table_row(ws)

    for row in range(11, last_table_row + 2):
        for col in range(1, 7):
            ws.cell(column=col,row=row).border = Border(top=thin,left=thin,right=thin,bottom=thin)

    ws.merge_cells(f'A{last_table_row + 1}:E{last_table_row + 1}')
    ws[f'A{last_table_row + 1}'].value = 'TỔNG THANH TOÁN'
    ws[f'F{last_table_row + 1}'].value = f'=SUM(F12:F{last_table_row})'
    ws[f'F{last_table_row + 1}'].number_format = '#,##0'

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    ws.row_dimensions[last_table_row + 1].height = 30
    ws[f'A{last_table_row + 1}'].alignment = Alignment(horizontal='center')

    ws[f'A{last_table_row + 1}'].font = Font(size=19,bold=True)
    ws[f'F{last_table_row + 1}'].font = Font(size=19,bold=True)
    print(customer_cell.value)

    wb.save(f'{out_dir}/{customer_cell.value.strip()}.xlsx')
    

def find_last_table_row(ws):
    for row in range(12, 1048576):
        if ws.cell(column=1,row=row).value == 'Khách hàng':
            return row - 4


if __name__ == "__main__":
    in_dir = "in"
    out_dir = "out"

    if not os.path.exists(out_dir):
        pathlib.Path(out_dir).mkdir(exist_ok=True)

    if len(sys.argv) == 2:
        in_dir = sys.argv[0]
        out_dir = sys.argv[1]

    format_files(in_dir, out_dir)